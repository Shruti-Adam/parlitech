from fastapi import APIRouter, HTTPException, Response
from fastapi.responses import FileResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
import csv
import io
import tempfile
import os
from datetime import datetime

router = APIRouter(prefix="/api/reports", tags=["reports"])

@router.post("/pdf")
async def generate_pdf(debate_data: dict):
    """Generate PDF report for debate"""
    try:
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
            filename = f"debate_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            doc = SimpleDocTemplate(tmp.name, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Custom title style
            title_style = ParagraphStyle(
                'CustomTitle', 
                parent=styles['Heading1'], 
                fontSize=24, 
                spaceAfter=30,
                textColor=colors.HexColor('#4f46e5')
            )
            
            # Title
            story.append(Paragraph("Parliamentary Debate Report", title_style))
            story.append(Spacer(1, 12))
            
            # Topic
            story.append(Paragraph(f"Topic: {debate_data.get('topic', 'N/A')}", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Date
            story.append(Paragraph(f"Date: {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Executive Summary
            story.append(Paragraph("Executive Summary", styles['Heading2']))
            summary_text = f"The parliamentary debate on '{debate_data.get('topic', 'the proposed legislation')}' has concluded. " \
                          f"Total speeches delivered: {len(debate_data.get('speeches', []))}. " \
                          f"Average sentiment: {((debate_data.get('sentiment_analysis', {}).get('average', 0)) * 100):.1f}%. " \
                          f"Public approval: {((debate_data.get('public_opinion', 0)) * 100):.1f}%."
            story.append(Paragraph(summary_text, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Voting Results
            story.append(Paragraph("Voting Results", styles['Heading2']))
            votes = debate_data.get('voting_results', {}).get('votes', {})
            total_votes = votes.get('for', 0) + votes.get('against', 0) + votes.get('abstain', 0)
            
            voting_data = [
                ['Vote Type', 'Count', 'Percentage'],
                ['For', str(votes.get('for', 0)), f"{(votes.get('for', 0)/total_votes*100) if total_votes > 0 else 0:.1f}%"],
                ['Against', str(votes.get('against', 0)), f"{(votes.get('against', 0)/total_votes*100) if total_votes > 0 else 0:.1f}%"],
                ['Abstain', str(votes.get('abstain', 0)), f"{(votes.get('abstain', 0)/total_votes*100) if total_votes > 0 else 0:.1f}%"],
            ]
            
            voting_table = Table(voting_data, colWidths=[100, 100, 100])
            voting_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4f46e5')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 12),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f3f4f6')),
            ]))
            story.append(voting_table)
            story.append(Spacer(1, 20))
            
            # Status
            status = debate_data.get('voting_results', {}).get('status', 'PENDING')
            status_style = ParagraphStyle(
                'StatusStyle',
                parent=styles['Normal'],
                fontSize=16,
                textColor=colors.HexColor('#10b981') if status == 'PASSED' else colors.HexColor('#ef4444')
            )
            story.append(Paragraph(f"Final Status: {status}", status_style))
            
            # Build PDF
            doc.build(story)
            
            # Return file
            return FileResponse(
                tmp.name,
                media_type='application/pdf',
                filename=filename,
                background=None
            )
    except Exception as e:
        print(f"PDF Generation Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/excel")
async def generate_excel(debate_data: dict):
    """Generate Excel report"""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            filename = f"debate_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            wb = Workbook()
            
            # Summary Sheet
            ws_summary = wb.active
            ws_summary.title = "Debate Summary"
            ws_summary.append(["Parameter", "Value"])
            ws_summary.append(["Topic", debate_data.get('topic', 'N/A')])
            ws_summary.append(["Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            ws_summary.append(["Status", debate_data.get('voting_results', {}).get('status', 'N/A')])
            ws_summary.append(["Total Speeches", len(debate_data.get('speeches', []))])
            ws_summary.append(["Average Sentiment", f"{((debate_data.get('sentiment_analysis', {}).get('average', 0)) * 100):.1f}%"])
            ws_summary.append(["Public Opinion", f"{((debate_data.get('public_opinion', 0)) * 100):.1f}%"])
            
            # Voting Results Sheet
            ws_voting = wb.create_sheet("Voting Results")
            ws_voting.append(["Party", "Seats", "For", "Against", "Abstain", "Support %"])
            
            party_breakdown = debate_data.get('voting_results', {}).get('party_breakdown', {})
            for party, data in party_breakdown.items():
                ws_voting.append([
                    party,
                    data.get('total_seats', 0),
                    data.get('for', 0),
                    data.get('against', 0),
                    data.get('abstain', 0),
                    data.get('support_percentage', 0)
                ])
            
            # Speeches Sheet
            ws_speeches = wb.create_sheet("Speeches")
            ws_speeches.append(["Speaker", "Role", "Party", "Sentiment", "Vote", "Speech Text"])
            
            for speech in debate_data.get('speeches', []):
                ws_speeches.append([
                    speech.get('agent_name', ''),
                    speech.get('agent_role', ''),
                    speech.get('party', ''),
                    speech.get('sentiment_score', 0),
                    speech.get('voting_tendency', ''),
                    speech.get('speech_text', '')[:500]  # Limit text length
                ])
            
            # Agent Performance Sheet
            ws_agents = wb.create_sheet("Agent Performance")
            ws_agents.append(["Agent", "Role", "Influence Score", "Speeches", "Avg Sentiment"])
            
            for agent in debate_data.get('agent_participation', []):
                ws_agents.append([
                    agent.get('name', ''),
                    agent.get('role', ''),
                    agent.get('influence_score', 0),
                    agent.get('speeches_given', 0),
                    agent.get('avg_sentiment', 0)
                ])
            
            wb.save(tmp.name)
            
            return FileResponse(
                tmp.name,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename=filename,
                background=None
            )
    except Exception as e:
        print(f"Excel Generation Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/csv")
async def generate_csv(debate_data: dict):
    """Generate CSV report"""
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['PARLITECH DEBATE REPORT'])
        writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
        writer.writerow([])
        
        # Basic Info
        writer.writerow(['Basic Information'])
        writer.writerow(['Topic', debate_data.get('topic', 'N/A')])
        writer.writerow(['Status', debate_data.get('voting_results', {}).get('status', 'N/A')])
        writer.writerow(['Total Speeches', len(debate_data.get('speeches', []))])
        writer.writerow(['Public Opinion', f"{((debate_data.get('public_opinion', 0)) * 100):.1f}%"])
        writer.writerow([])
        
        # Voting Results
        writer.writerow(['Voting Results'])
        writer.writerow(['Vote Type', 'Count', 'Percentage'])
        votes = debate_data.get('voting_results', {}).get('votes', {})
        total = votes.get('for', 0) + votes.get('against', 0) + votes.get('abstain', 0)
        writer.writerow(['For', votes.get('for', 0), f"{(votes.get('for', 0)/total*100) if total > 0 else 0:.1f}%"])
        writer.writerow(['Against', votes.get('against', 0), f"{(votes.get('against', 0)/total*100) if total > 0 else 0:.1f}%"])
        writer.writerow(['Abstain', votes.get('abstain', 0), f"{(votes.get('abstain', 0)/total*100) if total > 0 else 0:.1f}%"])
        writer.writerow([])
        
        # Party Breakdown
        writer.writerow(['Party-wise Voting'])
        writer.writerow(['Party', 'Seats', 'For', 'Against', 'Abstain', 'Support %'])
        for party, data in debate_data.get('voting_results', {}).get('party_breakdown', {}).items():
            writer.writerow([
                party,
                data.get('total_seats', 0),
                data.get('for', 0),
                data.get('against', 0),
                data.get('abstain', 0),
                data.get('support_percentage', 0)
            ])
        writer.writerow([])
        
        # Speeches Summary
        writer.writerow(['Speeches Summary'])
        writer.writerow(['Speaker', 'Role', 'Sentiment', 'Vote'])
        for speech in debate_data.get('speeches', []):
            writer.writerow([
                speech.get('agent_name', ''),
                speech.get('agent_role', ''),
                f"{((speech.get('sentiment_score', 0)) * 100):.1f}%",
                speech.get('voting_tendency', '')
            ])
        
        filename = f"debate_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return Response(
            content=output.getvalue().encode('utf-8-sig'),
            media_type='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        print(f"CSV Generation Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/research-paper")
async def generate_research_paper(debate_data: dict):
    """Generate research paper in academic format"""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
            filename = f"research_paper_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            doc = SimpleDocTemplate(tmp.name, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'ResearchTitle',
                parent=styles['Title'],
                fontSize=18,
                spaceAfter=20,
                alignment=1  # Center
            )
            story.append(Paragraph("AI-Powered Multi-Agent Parliamentary Intelligence System", title_style))
            story.append(Spacer(1, 10))
            
            # Authors
            author_style = ParagraphStyle('Author', parent=styles['Normal'], fontSize=12, alignment=1)
            story.append(Paragraph("ParliTech Research Team", author_style))
            story.append(Paragraph(f"Date: {datetime.now().strftime('%B %d, %Y')}", author_style))
            story.append(Spacer(1, 30))
            
            # Abstract
            story.append(Paragraph("ABSTRACT", styles['Heading2']))
            abstract_text = f"This research paper presents a comprehensive analysis of the parliamentary debate on '{debate_data.get('topic', 'proposed legislation')}' using the ParliTech AI-powered Multi-Agent Parliamentary Intelligence System. The system employs 15 distinct AI agents representing various political ideologies, simulating a realistic Lok Sabha debate environment. The debate concluded with a {debate_data.get('voting_results', {}).get('status', 'PENDING')} outcome, with {debate_data.get('voting_results', {}).get('votes', {}).get('for', 0)} votes in favor, {debate_data.get('voting_results', {}).get('votes', {}).get('against', 0)} against, and {debate_data.get('voting_results', {}).get('votes', {}).get('abstain', 0)} abstentions."
            story.append(Paragraph(abstract_text, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Introduction
            story.append(Paragraph("1. INTRODUCTION", styles['Heading2']))
            story.append(Paragraph("The Indian Parliament, particularly the Lok Sabha, represents one of the world's largest democratic deliberative bodies. Understanding the dynamics of parliamentary debate is crucial for political science research, policy analysis, and democratic governance studies. Traditional parliamentary analysis relies on manual observation and subjective interpretation. There is a need for an objective, data-driven approach to analyze debate dynamics, sentiment patterns, and voting behaviors.", styles['Normal']))
            story.append(Spacer(1, 15))
            
            # Methodology
            story.append(Paragraph("2. METHODOLOGY", styles['Heading2']))
            story.append(Paragraph("The ParliTech system comprises three primary layers: Multi-Agent Layer with 15 AI agents representing key parliamentary roles, each with distinct political ideology and memory system; Debate Orchestration Layer for sequential speech generation with real-time delivery and sentiment analysis; and Analytics and Visualization Layer with 3D Parliament chamber visualization and real-time analytics dashboard.", styles['Normal']))
            story.append(Spacer(1, 15))
            
            # Results
            story.append(Paragraph("3. RESULTS", styles['Heading2']))
            votes = debate_data.get('voting_results', {}).get('votes', {})
            results_text = f"The voting outcome showed {votes.get('for', 0)} votes in favor ({((votes.get('for', 0)/245)*100):.1f}%), {votes.get('against', 0)} against ({((votes.get('against', 0)/245)*100):.1f}%), and {votes.get('abstain', 0)} abstentions ({((votes.get('abstain', 0)/245)*100):.1f}%). The sentiment analysis revealed an average sentiment of {((debate_data.get('sentiment_analysis', {}).get('average', 0)) * 100):.1f}% with {debate_data.get('sentiment_analysis', {}).get('trend', 'neutral')} trend."
            story.append(Paragraph(results_text, styles['Normal']))
            story.append(Spacer(1, 15))
            
            # Conclusion
            story.append(Paragraph("4. CONCLUSION", styles['Heading2']))
            story.append(Paragraph("This research demonstrates that Multi-Agent AI Systems can effectively simulate and analyze parliamentary debates. The ParliTech system provides a scalable framework for political analysis, offering insights into voting pattern prediction, sentiment dynamics, coalition formation, and policy impact assessment.", styles['Normal']))
            story.append(Spacer(1, 15))
            
            # References
            story.append(Paragraph("5. REFERENCES", styles['Heading2']))
            references = [
                "Wooldridge, M. (2009). An Introduction to MultiAgent Systems. Wiley.",
                "Russell, S., & Norvig, P. (2020). Artificial Intelligence: A Modern Approach. Pearson.",
                "Liu, B. (2015). Sentiment Analysis: Mining Opinions, Sentiments, and Emotions.",
                "LeCun, Y., Bengio, Y., & Hinton, G. (2015). Deep learning. Nature."
            ]
            for ref in references:
                story.append(Paragraph(f"• {ref}", styles['Normal']))
            
            doc.build(story)
            
            return FileResponse(
                tmp.name,
                media_type='application/pdf',
                filename=filename,
                background=None
            )
    except Exception as e:
        print(f"Research Paper Generation Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))